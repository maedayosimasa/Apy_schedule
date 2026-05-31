"""日次 Excel バックアップ機能。

  ・1ヶ月分 = 1ブック  （例: 2026年05月.xlsx）
  ・1日分   = 1シート  （例: 05月31日）
  ・各シートはガントチャートと同じ2段構造で再現:
      - 上段（予定）: タスクパレット色 + 予定時間・内容テキスト
      - 下段（実績）: 実績時間テキスト（入力済=緑、未入力=薄灰）
      - 左固定列: 担当者（縦マージ）| 作業名（2段マージ）| 計(h)（2段マージ）
      - 日付ヘッダー: 今日=オレンジ、祝日=ピンク、土日=グレー
  ・バックアップ保存先は settings キー 'excel_backup_path' で管理。
    未設定時は DB ファイルと同階層の backup_excel/ フォルダ。
"""
import calendar
import json
from datetime import date as _date, timedelta
from itertools import groupby
from pathlib import Path

import database as db

try:
    from holidays import get_holiday as _get_holiday
except ImportError:
    def _get_holiday(d): return ''

_WEEKDAY   = ['月', '火', '水', '木', '金', '土', '日']
_STATUS_JP = {
    'planned':     '予定',
    'in_progress': '進行中',
    'completed':   '完了',
    'cancelled':   'キャンセル',
}
# ガントのタスクパレット（バー塗り色に対応する薄め版）
_TASK_PALETTE_HEX = [
    'BBDEFB',   # 水色
    'C8E6C9',   # 緑
    'FFF9C4',   # 黄
    'E1BEE7',   # 紫
    'FFCCBC',   # 橙
    'B2EBF2',   # ティール
]
# 上段予定行のテキスト色（タスクパレットに対応した濃い色）
_TASK_PALETTE_FG = [
    '1565C0',   # 青
    '2E7D32',   # 緑
    'E65100',   # 橙
    '6A1B9A',   # 紫
    'BF360C',   # 赤橙
    '006064',   # ティール
]


# ── 設定 ──────────────────────────────────────────────────────────────────────

def get_backup_dir() -> Path:
    v = db.get_setting('excel_backup_path', '')
    return Path(v) if v else Path(db.DB_PATH).parent / 'backup_excel'


def set_backup_dir(path: Path):
    db.set_setting('excel_backup_path', str(path))


def is_auto_backup_enabled() -> bool:
    return db.get_setting('excel_backup_enabled', '1') == '1'


def set_auto_backup_enabled(enabled: bool):
    db.set_setting('excel_backup_enabled', '1' if enabled else '0')


def get_last_backup_date() -> '_date | None':
    v = db.get_setting('excel_last_backup_date', '')
    try:
        return _date.fromisoformat(v) if v else None
    except ValueError:
        return None


def needs_startup_backup() -> bool:
    """起動時チェック: 自動バックアップが有効かつ昨日以前に未実行なら True。"""
    if not is_auto_backup_enabled():
        return False
    yesterday = _date.today() - timedelta(days=1)
    last = get_last_backup_date()
    return last is None or last < yesterday


def needs_startup_actual_backup() -> bool:
    """起動時チェック: 実績バックアップが今日まだ未実行なら True。"""
    if not is_auto_backup_enabled():
        return False
    v = db.get_setting('excel_last_actual_backup_date', '')
    try:
        last = _date.fromisoformat(v) if v else None
    except ValueError:
        last = None
    return last is None or last < _date.today()


# ── メインバックアップ処理 ────────────────────────────────────────────────────

def backup_day(target_date: _date | None = None,
               date_from:   _date | None = None,
               date_to:     _date | None = None) -> Path:
    """指定日のガントチャートを予定/実績2段構造で Excel に保存する。

    各タスク行が上段（予定）・下段（実績）の2行で構成され、
    ガントチャートのセルの上半分/下半分に対応する。

    Args:
        target_date : スナップショット日（シート名・ブック名に使用）。省略時は昨日。
        date_from   : 表示開始日。省略時は target_date の月初。
        date_to     : 表示終了日。省略時は target_date の月末。
    Returns:
        保存したブックの Path。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError(
            "openpyxl が見つかりません。\n"
            "コマンドプロンプトで以下を実行してください:\n\n"
            "  pip install openpyxl"
        )

    if target_date is None:
        target_date = _date.today() - timedelta(days=1)
    if date_from is None:
        date_from = target_date - timedelta(days=7)
    if date_to is None:
        date_to = target_date + timedelta(days=21)

    dates: list[_date] = []
    d = date_from
    while d <= date_to:
        dates.append(d); d += timedelta(days=1)
    if not dates:
        raise ValueError("date_from > date_to")

    ds_from = date_from.strftime('%Y-%m-%d')
    ds_to   = date_to.strftime('%Y-%m-%d')
    today   = _date.today()

    # ── ブック準備 ──
    save_dir = get_backup_dir()
    save_dir.mkdir(parents=True, exist_ok=True)
    book_name  = f"{target_date.year}年{target_date.month:02d}月.xlsx"
    book_path  = save_dir / book_name
    sheet_name = f"{target_date.month:02d}月{target_date.day:02d}日"

    if book_path.exists():
        wb = openpyxl.load_workbook(str(book_path))
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        _sort_sheets(wb)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

    # ── データ取得 ──
    schedules   = db.get_schedules(date_from=ds_from, date_to=ds_to)
    actuals     = db.get_actuals(date_from=ds_from, date_to=ds_to)
    all_workers = db.get_all_workers()
    all_tasks   = db.get_all_tasks()

    sched_lu  = {(s['worker_id'], s['task_id'], s['scheduled_date']): s
                 for s in schedules}
    actual_lu = {(a['worker_id'], a['task_id'], a['actual_date']): a
                 for a in actuals}

    w_map = {w['id']: w for w in all_workers}
    t_map = {t['id']: t for t in all_tasks}

    rows = _build_row_order(schedules, actuals, w_map, t_map)

    # ── 列インデックス ──
    COL_WORKER = 1
    COL_TASK   = 2
    COL_TOTAL  = 3
    COL_D0     = 4
    n_cols     = COL_D0 - 1 + len(dates)
    np_        = len(_TASK_PALETTE_HEX)

    # ── スタイルヘルパー ──
    def _f(bold=False, size=9, color='212121', italic=False):
        return Font(name='Yu Gothic UI', bold=bold, size=size,
                    color=color, italic=italic)
    def _fill(hex6):
        return PatternFill('solid', fgColor=hex6)

    _T  = Side(style='thin',   color='BDBDBD')
    _M  = Side(style='medium', color='455A64')
    _bdr      = Border(left=_T, right=_T, top=_T, bottom=_T)
    _bdr_bot  = Border(left=_T, right=_T, top=_T, bottom=_M)   # 実績行の下辺（担当者境界）
    _bdr_t    = Border(left=_T, right=_T, top=_M, bottom=_T)   # 予定行の上辺（担当者境界）
    _bdr_wl   = Border(left=_M, right=_T, top=_T, bottom=_T)
    _bdr_wtl  = Border(left=_M, right=_T, top=_M, bottom=_T)
    _bdr_wbl  = Border(left=_M, right=_T, top=_T, bottom=_M)   # 実績行下辺の担当者列

    c_aln  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_aln  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    cl_aln = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

    # ── 行1: タイトル ──
    _merge(ws, 1, 1, 1, n_cols)
    c = ws.cell(1, 1)
    wday = _WEEKDAY[target_date.weekday()]
    c.value = (f"ガントチャート  {target_date.year}年{target_date.month}月"
               f"{target_date.day}日（{wday}）スナップショット"
               f"　　表示期間: {date_from.month}/{date_from.day}"
               f" 〜 {date_to.month}/{date_to.day}")
    c.font = _f(bold=True, size=12, color='1A237E')
    c.fill = _fill('E8EAF6'); c.alignment = c_aln
    ws.row_dimensions[1].height = 22

    ws.row_dimensions[2].height = 4   # 空白行

    # ── 行3: 列ヘッダー ──
    HDR = 3
    for ci, (lbl, bg) in enumerate(
        [('担当者', '1565C0'), ('作業名', '1565C0'), ('計(h)', '1565C0')], 1
    ):
        c = ws.cell(HDR, ci, lbl)
        c.font = _f(bold=True, size=9, color='FFFFFF')
        c.fill = _fill(bg); c.alignment = c_aln; c.border = _bdr
    ws.row_dimensions[HDR].height = 26

    for di, d in enumerate(dates):
        ci      = COL_D0 + di
        holiday = _get_holiday(d)
        if d == today:              hbg, hfg = 'E65100', 'FFFFFF'
        elif holiday:               hbg, hfg = 'FFCDD2', 'C62828'
        elif d.weekday() >= 5:      hbg, hfg = 'BDBDBD', 'C62828'
        else:                       hbg, hfg = 'E3F2FD', '212121'
        c = ws.cell(HDR, ci, f"{d.month}/{d.day}\n({_WEEKDAY[d.weekday()]})")
        c.font = _f(bold=True, size=8, color=hfg)
        c.fill = _fill(hbg); c.alignment = c_aln; c.border = _bdr
        ws.column_dimensions[_col_letter(ci)].width = 6.5

    # ── データ行（タスク1件 = 予定行 + 実績行の2段） ──
    DATA_START = HDR + 1
    xrow = DATA_START

    # タスクカラー割当（ガントと同じアルゴリズム）
    task_color: dict[int, int] = {}
    row_cidx: list[int] = []
    for ri, r in enumerate(rows):
        tid = r['tid']
        if tid in task_color:
            row_cidx.append(task_color[tid])
        else:
            avoid = {row_cidx[ri - 1]} if ri > 0 and rows[ri-1]['tid'] != tid else set()
            for ci_try in range(np_):
                if ci_try not in avoid:
                    task_color[tid] = ci_try; row_cidx.append(ci_try); break
            else:
                task_color[tid] = 0; row_cidx.append(0)

    if not rows:
        _merge(ws, xrow, 1, xrow, n_cols)
        c = ws.cell(xrow, 1, "この期間のデータはありません")
        c.font = _f(size=10, color='9E9E9E', italic=True); c.alignment = c_aln
        _note_start = xrow + 2
    else:
        for wname, grp in groupby(rows, key=lambda r: r['wname']):
            group_rows = list(grp)
            grp_size   = len(group_rows)
            grp_start  = xrow

            for k, r in enumerate(group_rows):
                cidx        = row_cidx[rows.index(r)]
                wid, tid    = r['wid'], r['tid']
                is_grp_top  = (k == 0)
                is_grp_bot  = (k == grp_size - 1)

                srow = xrow       # 予定行
                arow = xrow + 1   # 実績行

                # 計(h) = 期間内の予定合計
                total_sched = sum(
                    sched_lu.get((wid, tid, d.strftime('%Y-%m-%d')), {})
                              .get('scheduled_hours', 0)
                    for d in dates
                )

                bar_bg  = _TASK_PALETTE_HEX[cidx % np_]
                bar_fg  = _TASK_PALETTE_FG[cidx % np_]
                alt_bg  = 'F5F5F5' if k % 2 else 'FFFFFF'

                # ── 担当者列（最初のタスクのみ書き込み、後でマージ） ──
                if is_grp_top:
                    c = ws.cell(srow, COL_WORKER, wname)
                    c.font = _f(bold=True, size=9)
                    c.fill = _fill('E3F2FD'); c.alignment = c_aln
                    c.border = _bdr_wtl    # 上辺太（担当者グループ境界）

                # ── 作業名（予定行・実績行をマージ） ──
                _merge(ws, srow, COL_TASK, arow, COL_TASK)
                c = ws.cell(srow, COL_TASK, r['tname'])
                c.font = _f(size=9); c.fill = _fill(alt_bg); c.alignment = l_aln
                c.border = _bdr_t if is_grp_top else _bdr

                # ── 計(h)（予定行・実績行をマージ） ──
                _merge(ws, srow, COL_TOTAL, arow, COL_TOTAL)
                c = ws.cell(srow, COL_TOTAL,
                            round(total_sched, 1) if total_sched > 0 else '')
                c.font = _f(bold=True, size=9)
                c.fill = _fill('E8F4FD' if k % 2 else 'FFFFFF')
                c.alignment = c_aln
                c.border = _bdr_t if is_grp_top else _bdr

                # ── 日付列 ──
                for di, d in enumerate(dates):
                    ci = COL_D0 + di
                    ds = d.strftime('%Y-%m-%d')
                    sched  = sched_lu.get((wid, tid, ds))
                    actual = actual_lu.get((wid, tid, ds))

                    holiday    = _get_holiday(d)
                    is_today_d = (d == today)
                    is_weekend = (d.weekday() >= 5)
                    is_past    = (d <= today)

                    # ── 予定行（上段）──────────────────────────────
                    if is_today_d:   sbg = 'FFFDE7'
                    elif holiday:    sbg = 'FFD9DC'
                    elif is_weekend: sbg = 'EEEEEE'
                    elif sched:      sbg = bar_bg
                    else:            sbg = alt_bg

                    # 予定行テキスト: 時間は非表示、連続スパンの全セルに作業項目を表示
                    sched_text = sched.get('note', '') if sched else ''

                    c = ws.cell(srow, ci, sched_text)
                    c.font      = _f(bold=bool(sched), size=8, color=bar_fg if sched else '757575')
                    c.fill      = _fill(sbg)
                    c.alignment = cl_aln if sched_text else c_aln
                    c.border    = _bdr_t if is_grp_top else _bdr

                    # ── 実績行（下段）──────────────────────────────
                    if is_today_d:   abg = 'FFFDE7'
                    elif holiday:    abg = 'FFD9DC'
                    elif is_weekend: abg = 'EEEEEE'
                    elif actual:     abg = 'E8F5E9'   # 実績入力済 = 薄緑
                    elif is_past:    abg = 'F5F5F5'   # 過去・未入力 = 薄灰
                    else:            abg = 'FFFFFF'

                    act_text = ''
                    if actual:
                        ah   = actual['actual_hours']
                        anote = actual.get('note', '')
                        parts = []
                        if ah > 0:
                            parts.append(f"{ah:.4g}h")
                        if anote:
                            parts.append(anote)
                        act_text = '\n'.join(parts)

                    c = ws.cell(arow, ci, act_text)
                    c.font      = _f(bold=bool(act_text), size=8,
                                     color='1B5E20' if act_text else '9E9E9E')
                    c.fill      = _fill(abg)
                    c.alignment = cl_aln if act_text else c_aln
                    # 実績行の下辺: 担当者グループの最後なら太線
                    c.border = _bdr_bot if is_grp_bot else _bdr

                ws.row_dimensions[srow].height = 20   # 予定行
                ws.row_dimensions[arow].height = 20   # 実績行（時間＋実施内容を表示）
                xrow += 2   # 2行分進める

            # 担当者セルをグループ全行分（grp_size×2）マージ
            _merge(ws, grp_start, COL_WORKER,
                   grp_start + grp_size * 2 - 1, COL_WORKER)
            c = ws.cell(grp_start, COL_WORKER)
            c.alignment = c_aln
            # マージ後に下辺のみ太線（グループ末尾）
            c.border = Border(left=_M, right=_T, top=_M, bottom=_M)

        # ── 合計行（予定/実績の2行） ──
        sum_srow = xrow
        sum_arow = xrow + 1
        SUM_BG = 'ECEFF1'

        for row_, label in [(sum_srow, '合計(予定)'), (sum_arow, '合計(実績)')]:
            _merge(ws, row_, COL_WORKER, row_, COL_TASK)
            c = ws.cell(row_, COL_WORKER, label)
            c.font = _f(bold=True, size=9); c.fill = _fill(SUM_BG)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = _bdr
            ws.cell(row_, COL_TASK).fill = _fill(SUM_BG)
            ws.cell(row_, COL_TASK).border = _bdr

        tot_sched  = sum(sched_lu.get((r['wid'], r['tid'], d.strftime('%Y-%m-%d')), {})
                         .get('scheduled_hours', 0)
                         for r in rows for d in dates)
        tot_actual = sum(actual_lu.get((r['wid'], r['tid'], d.strftime('%Y-%m-%d')), {})
                         .get('actual_hours', 0)
                         for r in rows for d in dates)

        c = ws.cell(sum_srow, COL_TOTAL,
                    round(tot_sched, 1) if tot_sched else '')
        c.font = _f(bold=True, size=9); c.fill = _fill(SUM_BG)
        c.alignment = c_aln; c.border = _bdr

        c = ws.cell(sum_arow, COL_TOTAL,
                    round(tot_actual, 1) if tot_actual else '')
        c.font = _f(bold=True, size=9, color='1B5E20')
        c.fill = _fill(SUM_BG); c.alignment = c_aln; c.border = _bdr

        for di, d in enumerate(dates):
            ci = COL_D0 + di
            ds = d.strftime('%Y-%m-%d')
            day_s = sum(sched_lu.get((r['wid'], r['tid'], ds), {})
                        .get('scheduled_hours', 0) for r in rows)
            day_a = sum(actual_lu.get((r['wid'], r['tid'], ds), {})
                        .get('actual_hours', 0) for r in rows)

            c = ws.cell(sum_srow, ci, f"{day_s:.4g}h" if day_s else '')
            c.font = _f(bold=True, size=8)
            c.fill = _fill(SUM_BG); c.alignment = c_aln; c.border = _bdr

            c = ws.cell(sum_arow, ci, f"{day_a:.4g}h" if day_a else '')
            c.font = _f(bold=True, size=8, color='1B5E20' if day_a else '212121')
            c.fill = _fill(SUM_BG); c.alignment = c_aln; c.border = _bdr

        ws.row_dimensions[sum_srow].height = 18
        ws.row_dimensions[sum_arow].height = 18
        _note_start = sum_arow + 2

    # ── 検討事項 ──
    gantt_note = db.get_setting('gantt_note', '')
    _N = Side(style='medium', color='F9A825')   # 黄色の枠線

    _merge(ws, _note_start, 1, _note_start, n_cols)
    c = ws.cell(_note_start, 1, '検討事項')
    c.font      = _f(bold=True, size=10, color='E65100')
    c.fill      = _fill('FFF9C4')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border    = Border(left=_N, right=_N, top=_N, bottom=Side(style='thin', color='F9A825'))
    ws.row_dimensions[_note_start].height = 22

    _note_body = _note_start + 1
    _merge(ws, _note_body, 1, _note_body, n_cols)
    c = ws.cell(_note_body, 1, gantt_note)
    c.font      = _f(size=10, color='333333')
    c.fill      = _fill('FFFDE7')
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    c.border    = Border(left=_N, right=_N, top=Side(style='thin', color='F9A825'), bottom=_N)
    ws.row_dimensions[_note_body].height = 80

    # ── 固定列の列幅 ──
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 7

    # ウィンドウ枠固定（ヘッダー行 + 左3列）
    ws.freeze_panes = ws.cell(DATA_START, COL_D0)

    wb.save(str(book_path))
    db.set_setting('excel_last_backup_date', target_date.isoformat())
    return book_path


def backup_month(year: int, month: int) -> Path:
    """指定月の各日スナップショットを一括作成する。"""
    _, last_day = calendar.monthrange(year, month)
    book_path = None
    for day in range(1, last_day + 1):
        d = _date(year, month, day)
        if d > _date.today():
            break
        book_path = backup_day(d)
    return book_path


def backup_actuals_month(year: int, month: int) -> 'Path | None':
    """指定月の実績データを 1ブック 1シートで保存する（毎回上書き）。

    ブック: 実績_{year}年{month:02d}月.xlsx
    シート: {month:02d}月実績（上書き）
    列構成: 担当者 | 作業名 | 実績日 | 曜日 | 実績時間(h) | 実施内容
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError(
            "openpyxl が見つかりません。pip install openpyxl"
        )

    _, last_day = calendar.monthrange(year, month)
    date_from = f"{year}-{month:02d}-01"
    date_to   = f"{year}-{month:02d}-{last_day:02d}"

    actuals = db.get_actuals(date_from=date_from, date_to=date_to)
    if not actuals:
        return None

    save_dir = get_backup_dir()
    save_dir.mkdir(parents=True, exist_ok=True)
    book_name  = f"実績_{year}年{month:02d}月.xlsx"
    book_path  = save_dir / book_name
    sheet_name = f"{month:02d}月実績"

    if book_path.exists():
        wb = openpyxl.load_workbook(str(book_path))
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

    # ── スタイルヘルパー ──
    def _f(bold=False, size=9, color='212121'):
        return Font(name='Yu Gothic UI', bold=bold, size=size, color=color)
    def _fill(hex6):
        return PatternFill('solid', fgColor=hex6)

    _T   = Side(style='thin',   color='BDBDBD')
    _M   = Side(style='medium', color='455A64')
    _bdr = Border(left=_T, right=_T, top=_T, bottom=_T)
    _bdr_hdr = Border(left=_M, right=_M, top=_M, bottom=_M)
    c_aln = Alignment(horizontal='center', vertical='center', wrap_text=True)
    l_aln = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    r_aln = Alignment(horizontal='right',  vertical='center')

    N_COLS = 6
    COL_WIDTHS = [12, 20, 10, 5, 13, 32]
    HEADERS    = ['担当者', '作業名', '実績日', '曜日', '実績時間(h)', '実施内容']

    # ── タイトル行 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
    c = ws.cell(1, 1, f"実績一覧　{year}年{month:02d}月")
    c.font = _f(bold=True, size=12, color='1A237E')
    c.fill = _fill('E8EAF6'); c.alignment = c_aln
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 4

    # ── ヘッダー行 ──
    HDR = 3
    for ci, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(HDR, ci, hdr)
        c.font = _f(bold=True, size=9, color='FFFFFF')
        c.fill = _fill('1565C0'); c.alignment = c_aln; c.border = _bdr_hdr
        ws.column_dimensions[_col_letter(ci)].width = w
    ws.row_dimensions[HDR].height = 24

    # ── データ行（日付 → 担当者 → 作業名順） ──
    sorted_actuals = sorted(
        actuals,
        key=lambda a: (a['actual_date'], a['worker_name'], a['task_title']),
    )

    prev_date = None
    for ri, a in enumerate(sorted_actuals, HDR + 1):
        d      = _date.fromisoformat(a['actual_date'])
        row_bg = 'F0F4F8' if ri % 2 == 0 else 'FFFFFF'

        # 日付が変わる行に上辺のアクセント線
        top_side = Side(style='medium', color='90A4AE') if d != prev_date else _T
        _bdr_row = Border(left=_T, right=_T, top=top_side, bottom=_T)
        prev_date = d

        holiday  = _get_holiday(d)
        is_we    = d.weekday() >= 5
        if holiday or is_we:
            row_bg = 'FFEEFF' if ri % 2 == 0 else 'FFF0FF'

        row_vals = [
            (a['worker_name'],                      c_aln),
            (a['task_title'],                       l_aln),
            (f"{d.month}/{d.day}",                  c_aln),
            (_WEEKDAY[d.weekday()],                 c_aln),
            (a['actual_hours'] if a['actual_hours'] > 0 else '', c_aln),
            (a.get('note', ''),                     l_aln),
        ]
        for ci, (val, aln) in enumerate(row_vals, 1):
            c = ws.cell(ri, ci, val)
            c.font = _f(size=9,
                        color='C62828' if (holiday or is_we) and ci == 4 else '212121')
            c.fill = _fill(row_bg); c.alignment = aln; c.border = _bdr_row
        ws.row_dimensions[ri].height = 16

    # ── 合計行 ──
    total_row = HDR + len(sorted_actuals) + 1
    total_h   = sum(a['actual_hours'] for a in sorted_actuals if a['actual_hours'])

    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    c = ws.cell(total_row, 1, f"合計  （{len(sorted_actuals)} 件）")
    c.font = _f(bold=True, size=9); c.fill = _fill('ECEFF1')
    c.alignment = r_aln; c.border = _bdr

    c = ws.cell(total_row, 5, round(total_h, 1) if total_h else '')
    c.font = _f(bold=True, size=9, color='1B5E20')
    c.fill = _fill('ECEFF1'); c.alignment = c_aln; c.border = _bdr

    ws.cell(total_row, 6).fill = _fill('ECEFF1')
    ws.cell(total_row, 6).border = _bdr
    ws.row_dimensions[total_row].height = 20

    ws.freeze_panes = ws.cell(HDR + 1, 1)

    wb.save(str(book_path))
    db.set_setting('excel_last_actual_backup_date', _date.today().isoformat())
    return book_path


# ── 行順序の構築 ──────────────────────────────────────────────────────────────

def _build_row_order(schedules, actuals, w_map, t_map) -> list[dict]:
    pair_info: dict[tuple, tuple] = {}
    for s in schedules:
        wid, tid = s['worker_id'], s['task_id']
        if wid in w_map and tid in t_map:
            pair_info[(wid, tid)] = (w_map[wid]['name'], t_map[tid]['title'])
    for a in actuals:
        wid, tid = a['worker_id'], a['task_id']
        if wid in w_map and tid in t_map:
            pair_info[(wid, tid)] = (w_map[wid]['name'], t_map[tid]['title'])
    if not pair_info:
        return []

    saved_order: list[tuple] = []
    try:
        v = db.get_setting('gantt_row_order', '')
        if v:
            saved_order = [tuple(p) for p in json.loads(v)]
    except Exception:
        pass

    existing     = [(w, t) for w, t in saved_order if (w, t) in pair_info]
    existing_set = set(existing)
    new_pairs    = sorted(
        [(w, t) for (w, t) in pair_info if (w, t) not in existing_set],
        key=lambda x: pair_info[x],
    )
    return [
        {'wid': w, 'tid': t,
         'wname': pair_info[(w, t)][0],
         'tname': pair_info[(w, t)][1]}
        for w, t in existing + new_pairs
        if (w, t) in pair_info
    ]


# ── ヘルパー ──────────────────────────────────────────────────────────────────

def _merge(ws, r1, c1, r2, c2):
    if r1 == r2 and c1 == c2:
        return
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _col_letter(n: int) -> str:
    result = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _sort_sheets(wb):
    names = wb.sheetnames
    try:
        def _key(n):
            if len(n) == 7 and n[2] == '月' and n[5] == '日':
                return (int(n[:2]), int(n[3:5]))
            return (99, 99)
        ordered = sorted(names, key=_key)
    except Exception:
        return
    for i, name in enumerate(ordered):
        cur = wb.sheetnames.index(name)
        if cur != i:
            wb.move_sheet(name, offset=i - cur)
